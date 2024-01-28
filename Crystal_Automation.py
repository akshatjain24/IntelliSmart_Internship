# importing all the required libraries
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sd = str(input("Enter Start Date (YYYY-MM-DD) format : "))
print("Note: The end date should be same as the end date of the file uploaded.")
ed = str(input("Enter End Date (YYYY-MM-DD) format : "))

# taking the range of date as input
start_date = datetime.strptime(sd, "%Y-%m-%d")
end_date = datetime.strptime(ed, "%Y-%m-%d")
D = 'D'
date_list = pd.date_range(start_date, end_date, freq=D)

# loading the required files in the memory
wb = load_workbook("compile.xlsx")
ws = wb.active
rd1 = pd.read_csv(f"SLA_LS_{ed.replace('-','_')}.csv", header=None)
rd2 = pd.read_csv(f"SLA_DP_{ed.replace('-','_')}.csv", header=None)

# Defining function to calculate the number of days in the given range
def numOfDays(date1, date2):
	if date2 > date1: 
		return (date2-date1).days
	else:
		return (date1-date2).days
	
col_src = 0
for i in rd1.iloc[0,0:]:
	if (str(i)[6:10] == sd[0:4] and str(i)[3:5] == sd[5:7] and str(i)[0:2] == sd[8:10]):
		break
	else:
		col_src += 1
xyz = col_src
l = []
j = 3
z = 2
b = 1

# Writing all meter number of ls  file
for i in rd1.iloc[1:, 0]:
	if i not in l:
		ws['A' + str(j)] = i
		j += 1
		l.append(i)
	b += 1
x = 1
for  w in rd2.iloc[1:, 0]:
	x+=1

# writing the data of ls file to the corresponding meter numbers 
for i in date_list:
	ws[get_column_letter(z) + '1'] = str(i)[0:10]
	ws.merge_cells(get_column_letter(z) + '1:' + get_column_letter(z+3)+'1')
	ws[get_column_letter(z) + '2'] = 'LS  8hr'
	ws[get_column_letter(z+1) + '2'] = 'LS  12hr'
	ws[get_column_letter(z+2) + '2'] = 'LS  24hr'
	ws[get_column_letter(z+3) + '2'] = 'DP  24hr'
	a = 3
	for w in rd1.iloc[1:, col_src]:
		ws[get_column_letter(z) + str(a)] = w
		a += 1
	a = 3
	for w in rd1.iloc[1:, col_src + 1]:
		ws[get_column_letter(z+1) + str(a)] = w
		a += 1
	a = 3
	for w in rd1.iloc[1:, col_src + 2]:
		ws[get_column_letter(z+2) + str(a)] = w
		a += 1
	a = 3
	for  w in rd2.iloc[1:, col_src + 2]:
		ws[get_column_letter(z+3) + str(a)] = f"=IFERROR(VLOOKUP(A{a}, SLA_DP_{str(ed)[0:10].replace('-','_')}.csv!$A$2:${get_column_letter(col_src+3)}${x-1},{col_src + 3},0),\"\")"
		a += 1
	z += 4
	col_src += 3

# Adding meter numbers and corresponding data of the DP file
z = 2
for i in rd2.iloc[1:,0]:
	if i not in l:
		ws['A' + str(j)] = i
		j += 1
		l.append(i)	
		for k in range(1, numOfDays(start_date,end_date)+1):
			ws[get_column_letter(5 + (k-1)*4)+ str(j)] = f"=IFERROR(VLOOKUP(A{j}, SLA_DP_{str(ed)[0:10].replace('-','_')}.csv!$A:${get_column_letter(col_src+1)},{xyz+3 + (k-1)*3},0),\"\")"			
	z+=1

# Flushing all the changes done on the file in the memory to the disc
print("All data compiled !!!")
wb.save("compile.xlsx")