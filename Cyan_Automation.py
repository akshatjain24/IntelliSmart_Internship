# importing all the required libraries
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd 
from datetime import datetime

# loading the master file in the memory
wbx = load_workbook(r'C:\Users\hp\Desktop\x_output\master.xlsx')
wsx = wbx.active

# taking the range of date as input
sd = input("Enter Start Date (YYYY-MM-DD) format : ")
ed = input("Enter End Date (YYYY-MM-DD) format : ")
start_date = datetime.strptime(sd, "%Y-%m-%d")
end_date = datetime.strptime(ed, "%Y-%m-%d")
D = 'D'
date_list = pd.date_range(start_date, end_date, freq=D)

# Defining function to calculate the number of days in the given range
def numOfDays(date1, date2):
	if date2 > date1: 
		return (date2-date1).days
	else:
		return (date1-date2).days

# Writing the data in the master file
l = []
col = 2
j = 3
x = 1 
for z in range(numOfDays(start_date, end_date) + 1): 
    print(f"Compiling day {x} of {numOfDays(start_date, end_date) + 1} ......")
    wsx.merge_cells(get_column_letter(col)+'1:'+ get_column_letter(col+3)+'1')
    wsx[get_column_letter(col) + '1'] = str(date_list[z])[0:10]
    wsx[get_column_letter(col) + '2'] = "LS  8hr"
    wsx[get_column_letter(col+1) + '2'] = "LS  12hr"
    wsx[get_column_letter(col+2) + '2'] = "LS  1d"
    wsx[get_column_letter(col+3) + '2'] = "DP  1d"

    # Reading the input csv files
    cs1 = pd.read_csv(f'C:\\Users\\hp\\Desktop\\CYAN\\LS\\{str(date_list[z])[0:10]}\\mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-1d.csv') 
    cs2 = pd.read_csv(f'C:\\Users\\hp\\Desktop\\CYAN\\LS\\{str(date_list[z])[0:10]}\\mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-12h.csv')
    cs3 = pd.read_csv(f'C:\\Users\\hp\\Desktop\\CYAN\\LS\\{str(date_list[z])[0:10]}\\mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-8h.csv')
    cs4 = pd.read_csv(f'C:\\Users\\hp\\Desktop\\CYAN\\DP\\mpsr-dailyload-{str(date_list[z])[0:10].replace("-","")}-1d.csv')

    # Converting the csv files to excel files to avoid windows security concerns while using the formulas in the master file
    cs1.to_excel(f'C:\\Users\\hp\\Desktop\\CYAN_Copy\\LS\\mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-1d.xlsx', sheet_name=f'mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-1d', index=False)
    cs2.to_excel(f'C:\\Users\\hp\\Desktop\\CYAN_Copy\\LS\\mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-12hr.xlsx', sheet_name=f'mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-12hr', index=False)
    cs3.to_excel(f'C:\\Users\\hp\\Desktop\\CYAN_Copy\\LS\\mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-8hr.xlsx', sheet_name=f'mpsr-blockload-{str(date_list[z])[0:10].replace("-","")}-8hr', index=False)
    cs4.to_excel(f'C:\\Users\\hp\\Desktop\\CYAN_Copy\\DP\\mpsr-dailyload-{str(date_list[z])[0:10].replace("-","")}-1d.xlsx', sheet_name=f'mpsr-dailyload-{str(date_list[z])[0:10].replace("-","")}-1d', index=False)
    
    # Writing unique Meter Numbers in the master file
    k = 3
    for i in cs1.iloc[2:,0]:
        k+=1
        if i not in l:
            wsx['A' + str(j)] = i
            j+=1
            l.append(i)

    # Writing the formulas the data cells to get the corresponding data of the Meter Numbers
    for i in range(3,j):
        wsx[get_column_letter(col) + str(i)] = f"=IFERROR(VLOOKUP(A{i},'C:\\Users\\hp\\Desktop\\CYAN_Copy\\LS\\[mpsr-blockload-{str(date_list[z])[0:10].replace('-','')}-8hr.xlsx]mpsr-blockload-{str(date_list[z])[0:10].replace('-','')}-8h'!$A$4:$B${k},2,0),\"\")"
        wsx[get_column_letter(col+1) + str(i)] = f"=IFERROR(VLOOKUP(A{i},'C:\\Users\\hp\\Desktop\\CYAN_Copy\\LS\\[mpsr-blockload-{str(date_list[z])[0:10].replace('-','')}-12hr.xlsx]mpsr-blockload-{str(date_list[z])[0:10].replace('-','')}-12h'!$A$4:$B${k},2,0),\"\")"
        wsx[get_column_letter(col+2) + str(i)] = f"=IFERROR(VLOOKUP(A{i},'C:\\Users\\hp\\Desktop\\CYAN_Copy\\LS\\[mpsr-blockload-{str(date_list[z])[0:10].replace('-','')}-1d.xlsx]mpsr-blockload-{str(date_list[z])[0:10].replace('-','')}-1d'!$A$4:$B${k},2,0),\"\")"
        wsx[get_column_letter(col+3) + str(i)] = f"=IFERROR(VLOOKUP(A{i},'C:\\Users\\hp\\Desktop\\CYAN_Copy\\DP\\[mpsr-dailyload-{str(date_list[z])[0:10].replace('-','')}-1d.xlsx]mpsr-dailyload-{str(date_list[z])[0:10].replace('-','')}-1d'!$A$4:$B${k},2,0),\"\")"
    col += 4
    x += 1

# Flushing all the changes done on the file in the memory to the disc
wbx.save(r'C:\Users\hp\Desktop\x_output\master.xlsx')
print("All Data compiled")